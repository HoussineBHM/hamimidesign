import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io, re

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Bank → Odoo",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

import streamlit.components.v1 as components

# ── Session state ──────────────────────────────────────────────────────────────
if 'page' not in st.session_state:
    st.session_state.page = 'intro'

# ── INTRO SCREEN ──────────────────────────────────────────────────────────────
if st.session_state.page == 'intro':

    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;1,300&family=Outfit:wght@200;300;400&display=swap');

#MainMenu,footer,header{visibility:hidden!important}
.block-container{padding:0!important;max-width:100%!important}
.stApp,[data-testid="stAppViewContainer"]{background:#080806!important}
section[data-testid="stSidebar"]{display:none!important}
[data-testid="stVerticalBlock"]{gap:0!important}

#sp-splash{
  position:fixed;inset:0;z-index:99999;background:#080806;
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  overflow:hidden;transition:opacity 1s ease,transform 1s ease;
}
#sp-splash.exit{opacity:0;transform:scale(1.05);pointer-events:none}
#sp-canvas{position:absolute;inset:0;width:100%;height:100%;z-index:0}

.sp-aurora{position:absolute;inset:0;z-index:1;pointer-events:none;overflow:hidden}
.sp-aurora::before{content:'';position:absolute;width:70vw;height:70vw;border-radius:50%;background:radial-gradient(ellipse,rgba(201,169,110,.07) 0%,rgba(184,115,74,.04) 40%,transparent 70%);top:-20%;left:-10%;animation:auroraA 14s ease-in-out infinite alternate}
.sp-aurora::after{content:'';position:absolute;width:60vw;height:60vw;border-radius:50%;background:radial-gradient(ellipse,rgba(184,115,74,.08) 0%,rgba(201,169,110,.03) 40%,transparent 70%);bottom:-20%;right:-10%;animation:auroraB 17s ease-in-out infinite alternate}
@keyframes auroraA{from{transform:translate(0,0) scale(1)}to{transform:translate(8%,6%) scale(1.1)}}
@keyframes auroraB{from{transform:translate(0,0) scale(1)}to{transform:translate(-8%,-6%) scale(1.12)}}

.sp-tile{position:absolute;inset:0;z-index:2;opacity:.28;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='60' height='60'%3E%3Crect x='1' y='1' width='58' height='58' fill='none' stroke='rgba(201,169,110,0.08)' stroke-width='1'/%3E%3Ccircle cx='30' cy='30' r='2' fill='rgba(201,169,110,0.06)'/%3E%3C/svg%3E");animation:tileDrift 40s linear infinite}
@keyframes tileDrift{to{background-position:60px 60px}}

.sp-scan{position:absolute;left:0;right:0;height:1px;z-index:3;pointer-events:none;background:linear-gradient(90deg,transparent,rgba(201,169,110,.2),rgba(201,169,110,.5),rgba(201,169,110,.2),transparent);box-shadow:0 0 12px rgba(201,169,110,.18);animation:scanMove 5s linear infinite}
@keyframes scanMove{0%{top:-1px;opacity:0}3%{opacity:1}97%{opacity:.4}100%{top:100%;opacity:0}}

.sp-corner{position:absolute;width:70px;height:70px}
.sp-corner::before,.sp-corner::after{content:'';position:absolute;background:#c9a96e;opacity:.28}
.sp-corner::before{width:100%;height:1px;top:0;left:0}
.sp-corner::after{width:1px;height:100%;top:0;left:0}
.spc-tl{top:24px;left:24px}.spc-tr{top:24px;right:24px;transform:scaleX(-1)}
.spc-bl{bottom:24px;left:24px;transform:scaleY(-1)}.spc-br{bottom:24px;right:24px;transform:scale(-1,-1)}

.sp-content{position:relative;z-index:10;display:flex;flex-direction:column;align-items:center;opacity:0;animation:spIn 1s cubic-bezier(.16,1,.3,1) .3s forwards;pointer-events:none}
@keyframes spIn{from{opacity:0;transform:translateY(22px)}to{opacity:1;transform:none}}

.sp-rings{position:relative;width:190px;height:190px;margin-bottom:26px;display:flex;align-items:center;justify-content:center}
.sp-ring{position:absolute;border-radius:50%;border:1px solid transparent}
.sp-ring-1{width:186px;height:186px;border-color:rgba(201,169,110,.09);border-top-color:rgba(201,169,110,.85);border-right-color:rgba(201,169,110,.35);animation:rA 4s linear infinite}
.sp-ring-2{width:142px;height:142px;border-color:rgba(184,115,74,.09);border-bottom-color:rgba(184,115,74,.88);border-left-color:rgba(184,115,74,.35);animation:rB 2.8s linear infinite}
.sp-ring-3{width:100px;height:100px;border-color:rgba(245,240,232,.06);border-top-color:rgba(245,240,232,.55);border-left-color:rgba(245,240,232,.22);animation:rA 2s linear infinite}
@keyframes rA{to{transform:rotate(360deg)}}@keyframes rB{to{transform:rotate(-360deg)}}
.sp-ring-dot{position:absolute;width:6px;height:6px;border-radius:50%;top:50%;left:50%}
.sp-ring-1 .sp-ring-dot{background:#c9a96e;box-shadow:0 0 8px rgba(201,169,110,.8);margin:-3px 0 0 90px}
.sp-ring-2 .sp-ring-dot{background:#b8734a;box-shadow:0 0 8px rgba(184,115,74,.8);margin:68px 0 0 -3px}
.sp-ring-3 .sp-ring-dot{background:#f5f0e8;box-shadow:0 0 6px rgba(245,240,232,.6);margin:-3px 0 0 -53px}

.sp-logo-center{position:absolute;z-index:5;display:flex;flex-direction:column;align-items:center;gap:2px}
.sp-h-logo{font-family:'Cormorant Garamond',serif;font-size:40px;font-weight:300;letter-spacing:.12em;color:#f5f0e8;line-height:1}
.sp-h-logo b{color:#c9a96e;font-weight:300}
.sp-h-sub{font-family:'Outfit',sans-serif;font-size:7px;letter-spacing:.32em;text-transform:uppercase;color:rgba(201,169,110,.45)}

.sp-title{font-family:'Cormorant Garamond',serif;font-size:15px;font-weight:300;letter-spacing:.32em;text-transform:uppercase;color:rgba(245,240,232,.75);margin-bottom:4px}
.sp-divider{width:0;height:1px;margin:10px 0 14px;background:linear-gradient(90deg,transparent,rgba(201,169,110,.5) 30%,rgba(184,115,74,.5) 70%,transparent);animation:divLine 1s ease .7s forwards}
@keyframes divLine{to{width:300px}}
.sp-tagline{font-family:'Outfit',sans-serif;font-size:9px;letter-spacing:.28em;text-transform:uppercase;color:rgba(140,132,120,.6);font-weight:300;animation:fadeU .8s ease .9s both}

.sp-counters{display:flex;gap:20px;margin:18px 0 22px;animation:fadeU .8s ease 1.1s both}
.sp-counter{display:flex;flex-direction:column;align-items:center;gap:3px;padding:10px 16px;border-radius:4px;background:rgba(201,169,110,.04);border:1px solid rgba(201,169,110,.1);min-width:80px}
.sp-counter-val{font-family:'Cormorant Garamond',serif;font-size:22px;font-weight:300;color:#c9a96e;line-height:1}
.sp-counter-lbl{font-family:'Outfit',sans-serif;font-size:8px;letter-spacing:.18em;text-transform:uppercase;color:rgba(140,132,120,.55);font-weight:300}

.sp-enter-wrap{margin-top:20px;display:flex;flex-direction:column;align-items:center;gap:7px;animation:fadeU .8s ease 1.4s both}
.sp-enter-visual{display:flex;align-items:center;gap:10px;padding:11px 30px;border-radius:2px;background:rgba(201,169,110,.05);border:1px solid rgba(201,169,110,.22);font-family:'Outfit',sans-serif;font-size:10px;letter-spacing:.28em;text-transform:uppercase;color:rgba(201,169,110,.7);font-weight:300;white-space:nowrap;animation:entGlow 3s ease-in-out 2s infinite}
@keyframes entGlow{0%,100%{border-color:rgba(201,169,110,.22);box-shadow:none}50%{border-color:rgba(201,169,110,.55);box-shadow:0 0 18px rgba(201,169,110,.1)}}
.sp-arr{display:inline-block;animation:arrSlide 1.3s ease-in-out infinite}
@keyframes arrSlide{0%,100%{transform:translateX(0)}50%{transform:translateX(4px)}}
.sp-hint{font-family:'Outfit',sans-serif;font-size:8px;color:rgba(140,132,120,.28);letter-spacing:.14em}

.sp-statusbar{position:absolute;bottom:0;left:0;right:0;padding:12px 24px;display:flex;justify-content:space-between;align-items:center;z-index:10}
.sp-sb-txt{font-family:'Outfit',sans-serif;font-size:8px;letter-spacing:.18em;text-transform:uppercase;color:rgba(140,132,120,.25)}
.sp-sb-bar{flex:1;max-width:220px;height:1px;margin:0 14px;border-radius:2px;background:rgba(201,169,110,.06);overflow:hidden}
.sp-sb-fill{height:100%;width:0;border-radius:2px;background:linear-gradient(90deg,#c9a96e,#b8734a);box-shadow:0 0 5px rgba(201,169,110,.35);animation:barFill 2.5s cubic-bezier(.4,0,.2,1) .5s forwards}
@keyframes barFill{0%{width:0}60%{width:75%}85%{width:90%}100%{width:100%}}
@keyframes fadeU{from{opacity:0;transform:translateY(7px)}to{opacity:1;transform:none}}

/* Invisible full-screen Streamlit button overlay */
div[data-testid="stButton"] > button {
  position:fixed!important;top:0!important;left:0!important;
  width:100vw!important;height:100vh!important;
  background:transparent!important;border:none!important;
  cursor:pointer!important;z-index:999999!important;
  opacity:0!important;padding:0!important;margin:0!important;border-radius:0!important;
}
</style>

<div id="sp-splash">
  <canvas id="sp-canvas"></canvas>
  <div class="sp-aurora"></div><div class="sp-tile"></div><div class="sp-scan"></div>
  <div class="sp-corner spc-tl"></div><div class="sp-corner spc-tr"></div>
  <div class="sp-corner spc-bl"></div><div class="sp-corner spc-br"></div>
  <div class="sp-content">
    <div class="sp-rings">
      <div class="sp-ring sp-ring-1"><div class="sp-ring-dot"></div></div>
      <div class="sp-ring sp-ring-2"><div class="sp-ring-dot"></div></div>
      <div class="sp-ring sp-ring-3"><div class="sp-ring-dot"></div></div>
      <div class="sp-logo-center">
        <div class="sp-h-logo">H<b>A</b>MIMI</div>
        <div class="sp-h-sub">Marrakech</div>
      </div>
    </div>
    <div class="sp-title">Finance · Import Tool</div>
    <div class="sp-divider"></div>
    <div class="sp-tagline">Bank Statements &#8594; Odoo &nbsp;·&nbsp; Powered by MSL-iTECH</div>
    <div class="sp-counters">
      <div class="sp-counter"><div class="sp-counter-val" id="sp-c1">0</div><div class="sp-counter-lbl">Formats</div></div>
      <div class="sp-counter"><div class="sp-counter-val" id="sp-c2">0</div><div class="sp-counter-lbl">Max rows</div></div>
      <div class="sp-counter"><div class="sp-counter-val" id="sp-c3">0</div><div class="sp-counter-lbl">Months</div></div>
    </div>
    <div class="sp-enter-wrap">
      <div class="sp-enter-visual"><span>Enter platform</span><span class="sp-arr">&#x2192;</span></div>
      <div class="sp-hint">Click anywhere to enter</div>
    </div>
  </div>
  <div class="sp-statusbar">
    <div class="sp-sb-txt">HAMIMI &nbsp;&#183;&nbsp; Finance</div>
    <div class="sp-sb-bar"><div class="sp-sb-fill"></div></div>
    <div class="sp-sb-txt">MSL-iTECH v2.0</div>
  </div>
</div>

<script>
(function(){
  var cv=document.getElementById('sp-canvas');
  if(!cv)return;
  var cx=cv.getContext('2d'),W,H,pts=[],alive=true,raf;
  function resize(){
    W=cv.width=window.innerWidth;H=cv.height=window.innerHeight;pts=[];
    var n=Math.min(Math.floor(W*H/14000),65);
    for(var i=0;i<n;i++)pts.push({x:Math.random()*W,y:Math.random()*H,vx:(Math.random()-.5)*.3,vy:(Math.random()-.5)*.3,r:Math.random()*1.4+.4,c:['201,169,110','184,115,74','245,240,232'][Math.floor(Math.random()*3)],a:.2+Math.random()*.4});
  }
  function frame(){
    if(!alive)return;
    cx.clearRect(0,0,W,H);
    for(var i=0;i<pts.length;i++){
      var p=pts[i];p.x+=p.vx;p.y+=p.vy;
      if(p.x<0||p.x>W)p.vx*=-1;if(p.y<0||p.y>H)p.vy*=-1;
      cx.beginPath();cx.arc(p.x,p.y,p.r,0,Math.PI*2);cx.fillStyle='rgba('+p.c+','+p.a+')';cx.fill();
      for(var j=i+1;j<pts.length;j++){var q=pts[j],d=Math.hypot(p.x-q.x,p.y-q.y);if(d<100){cx.beginPath();cx.moveTo(p.x,p.y);cx.lineTo(q.x,q.y);cx.strokeStyle='rgba(201,169,110,'+(1-d/100)*.055+')';cx.lineWidth=.5;cx.stroke();}}
    }
    raf=requestAnimationFrame(frame);
  }
  window.addEventListener('resize',resize);resize();frame();
  function cnt(id,target,suf,delay){
    setTimeout(function(){
      var el=document.getElementById(id);if(!el)return;
      var t0=null,dur=2000;
      (function step(ts){if(!t0)t0=ts;var p=Math.min((ts-t0)/dur,1),v=target*(1-Math.pow(1-p,3));el.textContent=Math.round(v)+suf;if(p<1)requestAnimationFrame(step);})(performance.now());
    },delay);
  }
  cnt('sp-c1',3,'',700);cnt('sp-c2',500,'+',900);cnt('sp-c3',12,'',700);
  window._splashStop=function(){alive=false;cancelAnimationFrame(raf);};
})();

document.addEventListener('mousedown',function(){
  var sp=document.getElementById('sp-splash');
  if(sp&&!sp.classList.contains('exit')){if(window._splashStop)window._splashStop();sp.classList.add('exit');}
});
document.addEventListener('keydown',function(e){
  if(e.key==='Enter'||e.key===' '){var sp=document.getElementById('sp-splash');if(sp&&!sp.classList.contains('exit')){if(window._splashStop)window._splashStop();sp.classList.add('exit');}}
});
</script>
""", unsafe_allow_html=True)

    # Invisible full-screen button — any click on the page triggers session state change
    if st.button("enter", key="intro_enter"):
        st.session_state.page = 'platform'
        st.rerun()

    st.stop()



# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');

*, *::before, *::after { box-sizing: border-box; }
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.stApp { background: #08090d; color: #c9cdd8; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 2.5rem 3rem; max-width: 1200px; }
section[data-testid="stSidebar"] { display: none; }

/* ── HERO ── */
.hero { margin-bottom: 3rem; }
.hero-eyebrow {
    font-family: 'DM Mono', monospace;
    font-size: 0.68rem;
    letter-spacing: 0.2em;
    color: #3ecf8e;
    text-transform: uppercase;
    margin-bottom: 0.6rem;
}
.hero-title {
    font-family: 'DM Mono', monospace;
    font-size: clamp(1.8rem, 3vw, 2.6rem);
    font-weight: 500;
    color: #f0f2f7;
    line-height: 1.1;
    margin: 0 0 0.8rem;
    letter-spacing: -0.03em;
}
.hero-title span { color: #3ecf8e; }
.hero-desc {
    font-size: 0.9rem;
    color: #5a6278;
    font-weight: 300;
    max-width: 520px;
    line-height: 1.6;
}
.hero-divider {
    height: 1px;
    background: linear-gradient(90deg, #1a1d26 0%, #2a2d3a 40%, #1a1d26 100%);
    margin: 2rem 0;
}

/* ── CARDS ── */
.card {
    background: #0e1018;
    border: 1px solid #1a1d26;
    border-radius: 12px;
    padding: 1.4rem 1.6rem;
    margin-bottom: 1rem;
    transition: border-color 0.2s;
}
.card:hover { border-color: #252836; }
.card-title {
    font-family: 'DM Mono', monospace;
    font-size: 0.68rem;
    letter-spacing: 0.15em;
    color: #3a3f52;
    text-transform: uppercase;
    margin-bottom: 1rem;
}

/* ── FILE ITEM ── */
.file-item {
    display: flex;
    align-items: center;
    gap: 0.9rem;
    padding: 0.75rem 1rem;
    background: #0b0d14;
    border: 1px solid #161824;
    border-radius: 8px;
    margin-bottom: 0.5rem;
}
.file-icon-wrap {
    width: 36px; height: 36px;
    background: #13151f;
    border: 1px solid #1e2130;
    border-radius: 8px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1rem; flex-shrink: 0;
}
.file-name { font-family: 'DM Mono', monospace; font-size: 0.78rem; color: #dde2f0; font-weight: 500; }
.file-sub  { font-size: 0.72rem; color: #3a3f52; margin-top: 1px; }
.file-right { margin-left: auto; }

/* ── BADGES ── */
.badge {
    font-family: 'DM Mono', monospace;
    font-size: 0.65rem;
    font-weight: 500;
    padding: 3px 10px;
    border-radius: 20px;
    letter-spacing: 0.06em;
    white-space: nowrap;
}
.badge-ok    { background: #0a2218; color: #3ecf8e; border: 1px solid #0e3325; }
.badge-err   { background: #1f0f0f; color: #f56565; border: 1px solid #2d1515; }
.badge-ready { background: #0a1628; color: #60a5fa; border: 1px solid #0e2040; }

/* ── FORMAT INFO ── */
.format-row {
    display: grid;
    grid-template-columns: 90px 1fr;
    gap: 0.3rem 1rem;
    font-size: 0.82rem;
    line-height: 1.9;
    align-items: center;
}
.format-key { font-family: 'DM Mono', monospace; color: #3ecf8e; font-size: 0.75rem; }
.format-val { color: #7a8099; }
.format-val em { color: #34d399; font-style: normal; }
.format-val s  { color: #ef4444; text-decoration-color: #ef4444; font-style: normal; text-decoration: line-through; }

/* ── SUPPORTED SOURCES ── */
.source-item {
    display: flex; align-items: center; gap: 0.6rem;
    padding: 0.45rem 0;
    border-bottom: 1px solid #111320;
    font-size: 0.8rem; color: #7a8099;
}
.source-item:last-child { border-bottom: none; }
.source-dot { width: 6px; height: 6px; border-radius: 50%; background: #3ecf8e; flex-shrink: 0; }

/* ── STATS ── */
.stats-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.8rem;
    margin: 1.5rem 0;
}
.stat-card {
    background: #0e1018;
    border: 1px solid #1a1d26;
    border-radius: 10px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.stat-num {
    font-family: 'DM Mono', monospace;
    font-size: 2rem;
    font-weight: 500;
    line-height: 1;
    margin-bottom: 0.4rem;
}
.stat-lbl {
    font-size: 0.65rem;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: #3a3f52;
}
.stat-total  { color: #a5b4fc; }
.stat-credit { color: #3ecf8e; }
.stat-debit  { color: #f87171; }
.stat-month  { color: #fbbf24; }

/* ── MAPPING CHIP ── */
.mapping-row {
    display: flex; flex-wrap: wrap; gap: 0.5rem;
    margin-top: 0.5rem;
}
.mapping-chip {
    font-family: 'DM Mono', monospace;
    font-size: 0.7rem;
    padding: 4px 10px;
    border-radius: 6px;
    border: 1px solid;
}
.chip-date   { background: #0d1a2e; color: #60a5fa; border-color: #1a3050; }
.chip-label  { background: #1a1220; color: #c084fc; border-color: #2a1a38; }
.chip-debit  { background: #1f0d0d; color: #f87171; border-color: #2d1515; }
.chip-credit { background: #071a12; color: #3ecf8e; border-color: #0d2e1f; }

/* ── DOWNLOAD CARD ── */
.dl-card {
    background: linear-gradient(135deg, #071a12 0%, #0a1f16 100%);
    border: 1px solid #0e3325;
    border-radius: 12px;
    padding: 1.6rem;
    text-align: center;
}
.dl-title { font-family: 'DM Mono', monospace; font-size: 0.75rem; color: #3ecf8e; margin-bottom: 1rem; letter-spacing: 0.1em; }
.dl-meta  { font-size: 0.72rem; color: #3a3f52; margin-top: 0.8rem; line-height: 1.8; }
.dl-month { color: #a78bfa; font-family: 'DM Mono', monospace; }

/* ── STREAMLIT OVERRIDES ── */
.stFileUploader > div {
    background: #0b0d14 !important;
    border: 1.5px dashed #1e2130 !important;
    border-radius: 10px !important;
    transition: border-color 0.2s !important;
}
.stFileUploader > div:hover { border-color: #3ecf8e !important; }
.stFileUploader label { color: #3a3f52 !important; font-size: 0.82rem !important; }
[data-testid="stFileUploaderDropzoneInstructions"] { color: #3a3f52 !important; }
[data-testid="stFileUploaderDropzoneInstructions"] svg { display:none; }

.stButton > button {
    background: #3ecf8e !important;
    color: #07120d !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'DM Mono', monospace !important;
    font-size: 0.85rem !important;
    font-weight: 500 !important;
    padding: 0.6rem 2rem !important;
    letter-spacing: 0.04em !important;
    width: 100% !important;
    transition: all 0.15s !important;
}
.stButton > button:hover {
    background: #2fb97e !important;
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 20px rgba(62,207,142,0.25) !important;
}
.stButton > button:active { transform: translateY(0) !important; }

.stDownloadButton > button {
    background: #3ecf8e !important;
    color: #07120d !important;
    border: none !important;
    border-radius: 8px !important;
    font-family: 'DM Mono', monospace !important;
    font-size: 0.85rem !important;
    font-weight: 500 !important;
    width: 100% !important;
    padding: 0.65rem !important;
    letter-spacing: 0.04em !important;
    transition: all 0.15s !important;
}
.stDownloadButton > button:hover {
    background: #2fb97e !important;
    box-shadow: 0 4px 20px rgba(62,207,142,0.3) !important;
}

div[data-testid="stDataFrame"] {
    border: 1px solid #1a1d26 !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}
.stTabs [data-baseweb="tab-list"] {
    background: #0b0d14;
    border-radius: 8px;
    padding: 3px;
    gap: 2px;
    border: 1px solid #1a1d26;
}
.stTabs [data-baseweb="tab"] {
    background: transparent;
    color: #3a3f52;
    font-family: 'DM Mono', monospace;
    font-size: 0.75rem;
    border-radius: 6px;
    letter-spacing: 0.05em;
}
.stTabs [aria-selected="true"] {
    background: #13151f !important;
    color: #dde2f0 !important;
}
.stSpinner > div { border-top-color: #3ecf8e !important; }
[data-testid="stNotification"] { border-radius: 8px !important; }
[data-testid="stAlertContainer"] { border-radius: 8px !important; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
MONTH_NAMES = {
    1:'janvier', 2:'février', 3:'mars', 4:'avril',
    5:'mai', 6:'juin', 7:'juillet', 8:'août',
    9:'septembre', 10:'octobre', 11:'novembre', 12:'décembre'
}

# Keywords that identify each column type — ORDER MATTERS (more specific first)
DATE_KEYWORDS   = ["date d'opération", "date operation", "date", "fecha"]
LABEL_KEYWORDS  = ["operation label", "libellé opération", "libelle operation",
                   "libellé", "libelle", "détail", "detail", "description",
                   "motif", "label", "narrative", "particulars", "memo"]
DEBIT_KEYWORDS  = ["débit", "debit", "sortie", "withdrawal", "payment", "montant débit"]
CREDIT_KEYWORDS = ["crédit", "credit", "entrée", "deposit", "montant crédit"]
IGNORE_KEYWORDS = ["solde", "balance", "date de valeur", "date valeur", "value date",
                   "référence", "reference", "ref", "devise", "currency"]

# ── Core logic ────────────────────────────────────────────────────────────────
def parse_amount(val):
    if pd.isna(val): return None
    cleaned = re.sub(r'[^\d.,-]', '', str(val)).replace(',', '.')
    # Handle "1.234.56" → take last dot as decimal
    parts = cleaned.split('.')
    if len(parts) > 2:
        cleaned = ''.join(parts[:-1]) + '.' + parts[-1]
    try: return float(cleaned)
    except: return None

def detect_date(val):
    s = str(val).strip()
    for fmt in ('%d.%m.%Y','%m/%d/%Y','%Y-%m-%d','%d/%m/%Y','%d-%m-%Y','%Y/%m/%d'):
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def score_column(col_name, keyword_list):
    """Return score 0-3 for how well col_name matches a keyword list."""
    c = str(col_name).strip().lower()
    for kw in keyword_list:
        if c == kw: return 3          # exact
        if c.startswith(kw): return 2 # starts with
        if kw in c: return 1          # contains
    return 0

def smart_map_columns(columns):
    """
    Universally map any column names to Date/Label/Debit/Credit.
    Uses scoring: ignores columns matching IGNORE_KEYWORDS,
    picks best match for each role.
    """
    # First filter out ignored columns
    candidates = []
    for c in columns:
        cl = str(c).strip().lower()
        if any(kw in cl for kw in IGNORE_KEYWORDS):
            continue
        candidates.append(c)

    roles = {'Date': None, 'Label': None, 'Debit': None, 'Credit': None}
    scores = {'Date': 0,   'Label': 0,    'Debit': 0,    'Credit': 0}

    for c in candidates:
        for role, keywords in [('Date', DATE_KEYWORDS), ('Label', LABEL_KEYWORDS),
                                ('Debit', DEBIT_KEYWORDS), ('Credit', CREDIT_KEYWORDS)]:
            s = score_column(c, keywords)
            if s > scores[role]:
                scores[role] = s
                roles[role] = c

    # Build col_map
    col_map = {}
    for role, col in roles.items():
        if col is not None:
            col_map[col] = role

    # Handle duplicate assignments: if two cols map to same role, keep higher score
    assigned = {}
    for col, role in col_map.items():
        if role not in assigned:
            assigned[role] = col
        else:
            # keep the one with higher score
            prev = assigned[role]
            if scores[role] < score_column(col, {
                'Date': DATE_KEYWORDS, 'Label': LABEL_KEYWORDS,
                'Debit': DEBIT_KEYWORDS, 'Credit': CREDIT_KEYWORDS
            }[role]):
                assigned[role] = col

    return {col: role for role, col in assigned.items()}

def find_header_row(raw_df, max_scan=15):
    """
    Scan first max_scan rows to find the header row.
    Looks for a row containing at least 2 of: date, label/detail/description, debit, credit keywords.
    Returns row index or None.
    """
    all_kws = DATE_KEYWORDS + LABEL_KEYWORDS + DEBIT_KEYWORDS + CREDIT_KEYWORDS
    best_row, best_score = None, 0

    for idx in range(min(max_scan, len(raw_df))):
        row_vals = [str(v).strip().lower() for v in raw_df.iloc[idx] if pd.notna(v) and str(v).strip()]
        score = 0
        for val in row_vals:
            if any(kw in val for kw in DATE_KEYWORDS): score += 2
            if any(kw in val for kw in LABEL_KEYWORDS): score += 2
            if any(kw in val for kw in DEBIT_KEYWORDS): score += 2
            if any(kw in val for kw in CREDIT_KEYWORDS): score += 2
        if score > best_score:
            best_score = score
            best_row = idx

    return best_row if best_score >= 4 else None

def clean_and_parse(df):
    """
    Given a dataframe with any column names:
    1. Smart-map columns
    2. Keep only Date/Label/Debit/Credit
    3. Filter out non-date rows (Total, blanks, metadata)
    4. Return clean_df and list of (date, label, amount) tuples
    """
    col_map = smart_map_columns(df.columns)
    df = df.rename(columns=col_map)

    keep = [c for c in ['Date', 'Label', 'Debit', 'Credit'] if c in df.columns]
    if 'Date' not in keep:
        return pd.DataFrame(), []

    df = df[keep].copy()

    # Filter: only rows with a valid parseable date
    mask = df['Date'].apply(lambda v: detect_date(str(v)) is not None)
    df = df[mask].copy()
    df['Date'] = df['Date'].apply(lambda v: detect_date(str(v)))

    # Drop rows with empty label (if label col exists)
    if 'Label' in df.columns:
        df = df[df['Label'].apply(lambda v: str(v).strip() not in ('nan', '', 'None'))].copy()

    df = df.reset_index(drop=True)

    # Build odoo rows
    rows = []
    for _, row in df.iterrows():
        date  = row['Date']
        label = str(row.get('Label', '')).strip()
        credit = parse_amount(row.get('Credit'))
        debit  = parse_amount(row.get('Debit'))
        if credit is not None and credit > 0:
            amount = credit
        elif debit is not None and debit > 0:
            amount = -debit
        else:
            amount = 0.0
        rows.append((date, label, amount))

    return df, rows

def parse_file(uploaded_file):
    """Universal file parser. Returns (all_rows, clean_df, mapping_info, errors)."""
    name = uploaded_file.name.lower()
    all_rows = []
    clean_frames = []
    mapping_info = {}
    errors = []

    try:
        if name.endswith('.csv'):
            content = uploaded_file.read()
            df = None
            for skip in range(8):
                try:
                    cand = pd.read_csv(io.BytesIO(content), skiprows=skip, encoding='utf-8-sig')
                    col_lower = [str(c).lower() for c in cand.columns]
                    if any('date' in c for c in col_lower):
                        df = cand
                        break
                except: pass

            if df is not None:
                clean_df, rows = clean_and_parse(df)
                if not clean_df.empty:
                    clean_frames.append(clean_df)
                    all_rows.extend(rows)
                    mapping_info = smart_map_columns(df.columns)

        elif name.endswith('.xls') or name.endswith('.xlsx'):
            engine = 'xlrd' if name.endswith('.xls') else 'openpyxl'
            xl = pd.ExcelFile(uploaded_file, engine=engine)

            for sheet_name in xl.sheet_names:
                raw = pd.read_excel(xl, sheet_name=sheet_name, header=None)

                header_row = find_header_row(raw)
                if header_row is None:
                    continue

                df = pd.read_excel(xl, sheet_name=sheet_name, header=header_row)
                clean_df, rows = clean_and_parse(df)

                if not clean_df.empty:
                    clean_frames.append(clean_df)
                    all_rows.extend(rows)
                    if not mapping_info:
                        mapping_info = smart_map_columns(df.columns)

    except Exception as e:
        errors.append(str(e))

    merged = pd.concat(clean_frames, ignore_index=True) if clean_frames else pd.DataFrame()
    return all_rows, merged, mapping_info, errors

def build_odoo_xlsx(rows):
    rows_by_month = {}
    month_order = []
    for date, label, amount in rows:
        mk = MONTH_NAMES[date.month]
        if mk not in rows_by_month:
            rows_by_month[mk] = []
            month_order.append(mk)
        rows_by_month[mk].append((date, label, amount))

    wb = Workbook()
    wb.remove(wb.active)
    hf = Font(name='Calibri', bold=True, size=11)
    df_font = Font(name='Calibri', size=11)
    header_fill = PatternFill('solid', fgColor='1a1d26')

    for mk in month_order:
        ws = wb.create_sheet(title=mk)
        for col, h in enumerate(['Date', 'Label', 'Amount'], 1):
            c = ws.cell(1, col, value=h)
            c.font = hf
        for i, (date, label, amount) in enumerate(rows_by_month[mk], start=2):
            c1 = ws.cell(i, 1, value=date);   c1.number_format = 'yyyy-mm-dd'; c1.font = df_font
            c2 = ws.cell(i, 2, value=label);  c2.font = df_font
            c3 = ws.cell(i, 3, value=amount); c3.number_format = '#,##0.00';   c3.font = df_font
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 72
        ws.column_dimensions['C'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, rows_by_month, month_order

# ── UI ────────────────────────────────────────────────────────────────────────

# Hero
st.markdown("""
<div class="hero">
  <div class="hero-eyebrow">● MSL-iTECH · Odoo Tools</div>
  <h1 class="hero-title">Bank <span>→</span> Odoo<br>Converter</h1>
  <p class="hero-desc">Importez n'importe quel relevé bancaire — le moteur détecte automatiquement les colonnes Date, Libellé, Débit, Crédit quelle que soit la structure du fichier.</p>
  <div class="hero-divider"></div>
</div>
""", unsafe_allow_html=True)

col_left, col_right = st.columns([1.1, 0.9], gap="large")

with col_left:
    # Upload
    st.markdown('<div class="card-title">① Upload</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "Déposez vos relevés ici",
        type=['csv', 'xls', 'xlsx'],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )
    st.markdown('<div style="font-size:0.7rem;color:#2a2d3a;margin-top:0.4rem;">CSV · XLS · XLSX · multi-fichiers · multi-feuilles</div>', unsafe_allow_html=True)

    if uploaded_files:
        st.markdown('<div class="card-title" style="margin-top:1.6rem;">② Fichiers chargés</div>', unsafe_allow_html=True)
        for f in uploaded_files:
            ext = f.name.split('.')[-1].upper()
            sz  = f"{round(f.size/1024,1)} KB"
            icon = "📊" if ext in ('XLS','XLSX') else "📄"
            st.markdown(f"""
            <div class="file-item">
              <div class="file-icon-wrap">{icon}</div>
              <div>
                <div class="file-name">{f.name}</div>
                <div class="file-sub">{ext} · {sz}</div>
              </div>
              <div class="file-right"><span class="badge badge-ready">PRÊT</span></div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<div style="margin-top:1.5rem;"></div>', unsafe_allow_html=True)
        convert_btn = st.button("⚡  Convertir en Odoo")
    else:
        convert_btn = False

with col_right:
    st.markdown('<div class="card-title">Format de sortie</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="card">
      <div class="format-row">
        <div class="format-key">Date</div>
        <div class="format-val"><code style="background:#0b0d14;padding:1px 6px;border-radius:4px;font-size:0.75rem;">yyyy-mm-dd</code></div>
        <div class="format-key">Label</div>
        <div class="format-val">Libellé de l'opération</div>
        <div class="format-key">Amount</div>
        <div class="format-val"><em>Crédit = positif (+)</em> · <s>Débit = négatif (−)</s></div>
        <div class="format-key">Feuilles</div>
        <div class="format-val">Une par mois — janvier, février…</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="card-title" style="margin-top:1.2rem;">Détection intelligente des colonnes</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="card">
      <div style="font-size:0.78rem;color:#5a6278;margin-bottom:0.8rem;line-height:1.6;">
        Le moteur scanne chaque fichier et associe automatiquement les colonnes selon leur nom,
        quelle que soit la langue ou la structure du relevé.
      </div>
      <div class="mapping-row">
        <span class="mapping-chip chip-date">date · date d'opération · fecha</span>
        <span class="mapping-chip chip-label">label · description · libellé · détail · motif</span>
        <span class="mapping-chip chip-debit">débit · debit · sortie · withdrawal</span>
        <span class="mapping-chip chip-credit">crédit · credit · entrée · deposit</span>
      </div>
      <div style="font-size:0.72rem;color:#2a2d3a;margin-top:0.8rem;">
        Colonnes ignorées automatiquement : Solde · Balance · Date de valeur · Référence…
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="card-title" style="margin-top:1.2rem;">Sources testées</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="card">
      <div class="source-item"><div class="source-dot"></div>CFG Bank — MAD (XLS multi-feuilles)</div>
      <div class="source-item"><div class="source-dot"></div>CFG Bank — MAD / Devise (XLS mono-feuille)</div>
      <div class="source-item"><div class="source-dot"></div>AJW — MAD / Devise (CSV)</div>
      <div class="source-item"><div class="source-dot" style="background:#5a6278;"></div>Tout relevé avec colonnes Date · Débit · Crédit</div>
    </div>
    """, unsafe_allow_html=True)

# ── PROCESS ───────────────────────────────────────────────────────────────────
if convert_btn and uploaded_files:
    st.markdown('<div class="hero-divider" style="margin:2rem 0;"></div>', unsafe_allow_html=True)

    all_rows   = []
    parse_log  = []

    with st.spinner("Analyse et conversion en cours…"):
        for f in uploaded_files:
            f.seek(0)
            rows, clean_df, mapping, errors = parse_file(f)
            all_rows.extend(rows)
            parse_log.append({'file': f.name, 'rows': len(rows),
                               'clean_df': clean_df, 'mapping': mapping, 'errors': errors})

    # Stats
    n_credit = sum(1 for _, _, a in all_rows if a > 0)
    n_debit  = sum(1 for _, _, a in all_rows if a < 0)
    months   = list(dict.fromkeys([MONTH_NAMES[r[0].month] for r in all_rows]))

    st.markdown(f"""
    <div class="stats-grid">
      <div class="stat-card">
        <div class="stat-num stat-total">{len(all_rows)}</div>
        <div class="stat-lbl">Transactions</div>
      </div>
      <div class="stat-card">
        <div class="stat-num stat-credit">{n_credit}</div>
        <div class="stat-lbl">Crédits</div>
      </div>
      <div class="stat-card">
        <div class="stat-num stat-debit">{n_debit}</div>
        <div class="stat-lbl">Débits</div>
      </div>
      <div class="stat-card">
        <div class="stat-num stat-month">{len(months)}</div>
        <div class="stat-lbl">Mois / Feuilles</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    res_col, dl_col = st.columns([2, 1], gap="large")

    with res_col:
        # Per-file result
        st.markdown('<div class="card-title">Résultat par fichier</div>', unsafe_allow_html=True)
        for log in parse_log:
            ok = log['rows'] > 0
            badge_cls = 'badge-ok' if ok else 'badge-err'
            badge_txt = f"{log['rows']} lignes" if ok else "ERREUR"
            icon = "✅" if ok else "❌"
            err_txt = " · ".join(log['errors']) if log['errors'] else "Aucune erreur"

            # Show detected mapping
            mapping_html = ""
            if log['mapping']:
                inv = {v: k for k, v in log['mapping'].items()}
                chips = {
                    'Date':   ('chip-date',   inv.get('Date',   '—')),
                    'Label':  ('chip-label',  inv.get('Label',  '—')),
                    'Debit':  ('chip-debit',  inv.get('Debit',  '—')),
                    'Credit': ('chip-credit', inv.get('Credit', '—')),
                }
                mapping_html = '<div class="mapping-row" style="margin-top:0.5rem;">' + ''.join(
                    f'<span class="mapping-chip {cls}" title="{role}">{col}</span>'
                    for role, (cls, col) in chips.items()
                ) + '</div>'

            st.markdown(f"""
            <div class="file-item" style="flex-direction:column;align-items:flex-start;padding:1rem;">
              <div style="display:flex;align-items:center;gap:0.8rem;width:100%;">
                <div style="font-size:1rem;">{icon}</div>
                <div style="flex:1;">
                  <div class="file-name">{log['file']}</div>
                  <div class="file-sub">{err_txt}</div>
                </div>
                <span class="badge {badge_cls}">{badge_txt}</span>
              </div>
              {mapping_html}
            </div>
            """, unsafe_allow_html=True)

        # Preview
        if all_rows:
            st.markdown('<div style="height:1.2rem;"></div>', unsafe_allow_html=True)
            st.markdown('<div class="card-title">Aperçu des données</div>', unsafe_allow_html=True)

            tab_clean, tab_odoo = st.tabs(["🧹  Fichier nettoyé", "✅  Format Odoo"])

            with tab_clean:
                all_clean = pd.concat(
                    [log['clean_df'] for log in parse_log if not log['clean_df'].empty],
                    ignore_index=True
                )
                disp = all_clean.copy()
                disp['Date'] = disp['Date'].apply(lambda d: d.strftime('%Y-%m-%d') if pd.notna(d) else '')
                st.dataframe(disp, use_container_width=True, hide_index=True,
                             height=min(480, 44 + len(disp)*36),
                             column_config={
                                 "Date":   st.column_config.TextColumn("Date", width="small"),
                                 "Label":  st.column_config.TextColumn("Label", width="large"),
                                 "Debit":  st.column_config.NumberColumn("Débit",  format="%.2f", width="small"),
                                 "Credit": st.column_config.NumberColumn("Crédit", format="%.2f", width="small"),
                             })

            with tab_odoo:
                _, rows_by_month, month_order = build_odoo_xlsx(all_rows)
                mtabs = st.tabs([m.capitalize() for m in month_order])
                for mtab, mk in zip(mtabs, month_order):
                    with mtab:
                        prows = rows_by_month[mk]
                        pdf = pd.DataFrame(prows, columns=['Date','Label','Amount'])
                        pdf['Date']   = pdf['Date'].dt.strftime('%Y-%m-%d')
                        pdf['Amount'] = pdf['Amount'].apply(lambda x: f"+{x:,.2f}" if x > 0 else f"{x:,.2f}")
                        st.dataframe(pdf, use_container_width=True, hide_index=True,
                                     height=min(480, 44 + len(prows)*36),
                                     column_config={
                                         "Date":   st.column_config.TextColumn("Date", width="small"),
                                         "Label":  st.column_config.TextColumn("Label", width="large"),
                                         "Amount": st.column_config.TextColumn("Amount", width="small"),
                                     })

    with dl_col:
        if all_rows:
            xlsx_buf, rows_by_month, month_order = build_odoo_xlsx(all_rows)
            base_names = [f.name.rsplit('.',1)[0] for f in uploaded_files]
            out_name = ("_".join(base_names[:2])+"_Odoo.xlsx") if len(base_names)<=2 else f"Bank_Odoo_{len(uploaded_files)}files.xlsx"

            month_lines = "".join(
                f'<div>↳ <span class="dl-month">{mk}</span> — {len(rows_by_month[mk])} lignes</div>'
                for mk in month_order
            )

            st.markdown(f"""
            <div class="dl-card">
              <div class="dl-title">⬇ FICHIER PRÊT</div>
            """, unsafe_allow_html=True)

            st.download_button(
                label="Télécharger le fichier Odoo",
                data=xlsx_buf,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.markdown(f"""
              <div class="dl-meta">
                <div style="color:#5a6278;margin-bottom:0.4rem;">{out_name}</div>
                {month_lines}
              </div>
            </div>
            <div style="margin-top:1rem;padding:1rem;background:#0b0d14;border:1px solid #1a1d26;border-radius:10px;">
              <div style="font-family:'DM Mono',monospace;font-size:0.65rem;color:#3a3f52;letter-spacing:0.1em;margin-bottom:0.6rem;">IMPORT ODOO</div>
              <div style="font-size:0.78rem;color:#5a6278;line-height:1.9;">
                Comptabilité<br>
                → Relevés bancaires<br>
                → <span style="color:#3ecf8e;">Importer</span><br>
                → Sélectionner le fichier
              </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.error("Aucune transaction extraite. Vérifiez le format des fichiers.")

elif not uploaded_files:
    st.markdown("""
    <div style="text-align:center;padding:4rem 2rem;color:#1e2130;">
      <div style="font-family:'DM Mono',monospace;font-size:3rem;margin-bottom:1rem;opacity:0.3;">🏦</div>
      <div style="font-family:'DM Mono',monospace;font-size:0.8rem;letter-spacing:0.1em;color:#2a2d3a;">
        DÉPOSEZ UN FICHIER POUR COMMENCER
      </div>
    </div>
    """, unsafe_allow_html=True)
